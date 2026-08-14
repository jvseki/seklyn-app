"""
Exportação profissional da ficha do aluno em Excel (.xlsx) — treinos
completos (dia a dia, exercício a exercício, série a série) e o
histórico de peso/meta, com a identidade visual roxa do Seklyn.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.aluno import Aluno

ROXO = "7C3AED"  # violet-600 — cor da marca Seklyn
ROXO_ESCURO = "4C1D95"
LILAS = "EDE9FE"
ZEBRA = "F7F5FF"
CINZA_TEXTO = "666666"

BORDA_FINA = Border(
    left=Side(style="thin", color="D9D3EC"),
    right=Side(style="thin", color="D9D3EC"),
    top=Side(style="thin", color="D9D3EC"),
    bottom=Side(style="thin", color="D9D3EC"),
)

ROTULO_DIA = {
    "segunda": "Segunda",
    "terca": "Terça",
    "quarta": "Quarta",
    "quinta": "Quinta",
    "sexta": "Sexta",
    "sabado": "Sábado",
    "domingo": "Domingo",
}


def _cabecalho(ws, titulo: str, subtitulo: str, ultima_coluna: str) -> int:
    """Escreve o título/subtítulo/crédito no topo da aba e devolve a próxima linha livre."""
    ws.merge_cells(f"A1:{ultima_coluna}1")
    t = ws["A1"]
    t.value = f"Seklyn — {titulo}"
    t.font = Font(name="Calibri", size=16, bold=True, color=ROXO_ESCURO)
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{ultima_coluna}2")
    s = ws["A2"]
    s.value = subtitulo
    s.font = Font(name="Calibri", size=11, color=CINZA_TEXTO)

    ws.merge_cells(f"A3:{ultima_coluna}3")
    c = ws["A3"]
    c.value = "Gerado pelo Seklyn — seklyn.com.br"
    c.font = Font(name="Calibri", size=9, italic=True, color="999999")

    return 5


def _linha_cabecalho_tabela(ws, linha: int, colunas: list[str]) -> None:
    fill = PatternFill("solid", fgColor=ROXO)
    fonte = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col, nome in enumerate(colunas, start=1):
        celula = ws.cell(linha, col, nome)
        celula.fill = fill
        celula.font = fonte
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = BORDA_FINA
    ws.row_dimensions[linha].height = 22


def build_aluno_xlsx(aluno: Aluno, avaliacoes: list) -> bytes:
    wb = Workbook()

    # ---------- Aba 1: Treinos ----------
    ws = wb.active
    ws.title = "Treinos"
    colunas_treino = ["Dia da semana", "Treino", "Exercício", "Série", "Repetições", "Carga", "Descanso"]
    ultima_col = get_column_letter(len(colunas_treino))
    linha = _cabecalho(ws, "Ficha de treino", f"Aluno: {aluno.nome}", ultima_col)
    _linha_cabecalho_tabela(ws, linha, colunas_treino)

    fonte_corpo = Font(name="Calibri", size=10)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    esquerda = Alignment(horizontal="left", vertical="center", wrap_text=True)
    zebra_fill = PatternFill("solid", fgColor=ZEBRA)

    linha_atual = linha + 1
    linha_par = False
    treinos_ordenados = sorted(aluno.treinos, key=lambda t: (t.ordem, t.nome))
    for treino in treinos_ordenados:
        exercicios_ordenados = sorted(treino.exercicios, key=lambda e: e.ordem)
        if not exercicios_ordenados:
            valores = [ROTULO_DIA.get(treino.dia_semana, treino.dia_semana or "—"), treino.nome, "—", "—", "—", "—", "—"]
            linhas_a_escrever = [valores]
        else:
            linhas_a_escrever = []
            for exercicio in exercicios_ordenados:
                series_ordenadas = sorted(exercicio.series, key=lambda s: s.ordem)
                if not series_ordenadas:
                    linhas_a_escrever.append(
                        [ROTULO_DIA.get(treino.dia_semana, treino.dia_semana or "—"), treino.nome, exercicio.nome, "—", "—", "—", "—"]
                    )
                for serie in series_ordenadas:
                    linhas_a_escrever.append(
                        [
                            ROTULO_DIA.get(treino.dia_semana, treino.dia_semana or "—"),
                            treino.nome,
                            exercicio.nome,
                            f"Série {serie.ordem + 1}",
                            serie.repeticoes_alvo,
                            serie.carga_alvo or "—",
                            serie.intervalo_descanso or "—",
                        ]
                    )

        linha_par = not linha_par
        for valores in linhas_a_escrever:
            for col, valor in enumerate(valores, start=1):
                celula = ws.cell(linha_atual, col, valor)
                celula.font = fonte_corpo
                celula.border = BORDA_FINA
                celula.alignment = centro if col in (1, 4) else esquerda
                if linha_par:
                    celula.fill = zebra_fill
            ws.row_dimensions[linha_atual].height = 18
            linha_atual += 1

    if linha_atual == linha + 1:
        ws.cell(linha_atual, 1, "Nenhum treino cadastrado ainda.").font = Font(name="Calibri", italic=True, color=CINZA_TEXTO)
        linha_atual += 1

    larguras = {"A": 14, "B": 22, "C": 26, "D": 12, "E": 14, "F": 14, "G": 14}
    for letra, largura in larguras.items():
        ws.column_dimensions[letra].width = largura
    ws.freeze_panes = f"A{linha + 1}"
    ws.auto_filter.ref = f"A{linha}:{ultima_col}{linha_atual - 1}"

    # ---------- Aba 2: Peso e meta ----------
    ws2 = wb.create_sheet("Peso e meta")
    colunas_peso = ["Data", "Peso (kg)", "Observações"]
    ultima_col2 = get_column_letter(len(colunas_peso))
    linha2 = _cabecalho(ws2, "Evolução de peso", f"Aluno: {aluno.nome}", ultima_col2)

    ws2.cell(linha2, 1, "Meta de peso").font = Font(name="Calibri", bold=True, color=ROXO_ESCURO)
    ws2.cell(linha2, 2, f"{aluno.peso_meta_kg} kg" if aluno.peso_meta_kg else "Não definida")
    linha2 += 2

    tabela_linha = linha2
    _linha_cabecalho_tabela(ws2, tabela_linha, colunas_peso)
    linha2 = tabela_linha + 1

    avaliacoes_ordenadas = sorted(avaliacoes, key=lambda a: a.data)
    for idx, avaliacao in enumerate(avaliacoes_ordenadas):
        valores = [avaliacao.data.strftime("%d/%m/%Y"), avaliacao.peso_kg, avaliacao.observacoes or "—"]
        for col, valor in enumerate(valores, start=1):
            celula = ws2.cell(linha2, col, valor)
            celula.font = fonte_corpo
            celula.border = BORDA_FINA
            celula.alignment = centro if col in (1, 2) else esquerda
            if idx % 2:
                celula.fill = zebra_fill
        ws2.row_dimensions[linha2].height = 18
        linha2 += 1

    if not avaliacoes_ordenadas:
        ws2.cell(linha2, 1, "Nenhum peso registrado ainda.").font = Font(name="Calibri", italic=True, color=CINZA_TEXTO)
        linha2 += 1

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 40
    ws2.freeze_panes = f"A{tabela_linha + 1}"
    ws2.auto_filter.ref = f"A{tabela_linha}:{ultima_col2}{linha2 - 1}"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
